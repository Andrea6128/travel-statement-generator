# travel statement generator
# this program generates random travel times and
# calculates corresponding values and fees
# 
# 'empty.xlsx' must exist and has to be formatted properly!
# 'mesta_input.xlsx' must exist!
#
# only 4-city model (2 routes per day) applied

from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
import random
import datetime
import sys
import decimal


def getPetrol():
    """ count petrol consumption price per 1km """

    petrolPrice = random.randint(120, 135) / 100  # random cena za litr
    flatRateConpensation = .193  # pausalni nahrada
    consumption = (22 / 100) + flatRateConpensation  # 22 litres per 100 km
    consumptionPrice = consumption * petrolPrice

    return [petrolPrice, consumptionPrice]


def randomTime(routeLength):
    """ generate random times and return start, end and total time """

    # convert routeLength minutes to a tuple (HOUR, MINUTE) and then to variables
    routeLengthHour = int(routeLength[:1])
    routeLengthMinute = int(routeLength[3:-3])

    # morning time generator
    randHour, randMinute = random.randint(0, 3), random.randint(0, 59)
    # salt = random.randint(0, 10)
    morningStartTime = datetime.datetime(1970, 1, 1, hour=6+randHour, minute=randMinute)
    morningAdd = datetime.timedelta(hours=routeLengthHour, minutes=routeLengthMinute) # + salt)
    morningEndTime = morningStartTime + morningAdd
    
    # evening time generator
    randHour, randMinute = random.randint(0, 1), random.randint(27, 59)
    salt = random.randint(0, 10)
    eveningStartTime = datetime.datetime(1970, 1, 1, hour=21+randHour, minute=randMinute)
    eveningAdd = datetime.timedelta(hours=routeLengthHour, minutes=routeLengthMinute) # + salt)
    eveningEndTime = eveningStartTime + eveningAdd

    # count total time (evening end - morning start)
    totalTime = str(eveningEndTime - morningStartTime)[:-3]
    
    # convert times to strings and cut unwanted chars
    morningStartTime = str(morningStartTime)[10:-3]
    morningEndTime = str(morningEndTime)[10:-3]
    eveningStartTime = str(eveningStartTime)[10:-3]
    eveningEndTime = str(eveningEndTime)[10:-3]

    return [morningStartTime, morningEndTime, eveningStartTime, eveningEndTime, totalTime]


def dayRoute():
    """ generate random start city """

    dayRouteResult = []
    ws = wb1.active  # set input excel active
    randomStartCity = random.randint(1, 72)  # select random city combination from the base city

    cityFrom = ws['A'+str(randomStartCity)].value
    cityTo = ws['B'+str(randomStartCity)].value
    distance = ws['C'+str(randomStartCity)].value
    travelTime = ws['D'+str(randomStartCity)].value

    dayRouteResult.append(cityFrom)
    dayRouteResult.append(cityTo)
    dayRouteResult.append(str(distance))
    dayRouteResult.append(str(travelTime))

    return dayRouteResult

def writeFooterValues():    
    # *** footer ***
    # write petrol column sum
    petrolValueList = []
    for row in range(5, 128, 2):
        petrolValueList.append(ws.cell(row=row, column=8).value)
    sumOfValueList = sum(petrolValueList)
    ws.cell(row=129, column=8, value=sumOfValueList)

    # write diets column sum
    dietValueList = []
    for row in range(5, 128, 2):
        dietValueList.append(ws.cell(row=row, column=9).value)
    sumOfValueList = sum(dietValueList)
    ws.cell(row=129, column=9, value=sumOfValueList)

    # write "together" fields sum
    togetherValueList = []
    for cell in range(8, 10):
        togetherValueList.append(float(ws.cell(row=129, column=cell).value))
    sumOfValueList = sum(togetherValueList)
    ws.cell(row=129, column=12, value=sumOfValueList)

    # write "overpay/underpay" field sum
    ws.cell(row=131, column=12, value=ws.cell(row=129, column=12).value)


def fillSheet(startRow, startColumn, startDate, numberOfDays):
    """ fill the worksheet with all data """

    for day in range(numberOfDays):
        getStartRoute = dayRoute()  # get dayRoute function result [start city, end city, km, time]
        getRandomTime = randomTime(getStartRoute[3])  # get start hours, end hours and travel time
                                                      # [start time, end time, start time2, end time2, total time]

        # set diets
        # if 5-12, set to 5.1
        if int(getRandomTime[4][:2]) >= 5 and int(getRandomTime[4][:2]) < 12:
            diets = round(decimal.Decimal(5.1), 2)
        #if 12-18, set to 7.6
        if int(getRandomTime[4][:2]) >= 12 and int(getRandomTime[4][:2]) < 18:
            diets = round(decimal.Decimal(7.6), 2)
        # if 18 and more, set to 11.6
        if int(getRandomTime[4][:2]) >= 18:
            diets = round(decimal.Decimal(11.6), 2)

        for event in range(4):
            # way there ->
            if event == 0:
                startDateStr = startDate.strftime('%d.%m.%Y')  # starting date (usualy 1.1.2020)
                ws.cell(row=startRow, column=startColumn, value=startDateStr)  # fill 1st cell with date
                ws.cell(row=startRow, column=startColumn+1, value="odchod")  # fill odchod
                ws.cell(row=startRow, column=startColumn+2, value=getStartRoute[0])  # fill start city
                ws.cell(row=startRow, column=startColumn+3, value=getRandomTime[0]).alignment = Alignment(vertical="center", horizontal="center")  # morning start time
                ws.cell(row=startRow, column=startColumn+4, value="AUS").alignment = Alignment(vertical="center", horizontal="center")  # set to AUS
                ws.cell(row=startRow, column=startColumn+5, value=getStartRoute[2]).alignment = Alignment(vertical="center", horizontal="center")  # km

                petrolPrice = getPetrol()  # get random petrol price
                petrolPrice = float(round(decimal.Decimal(petrolPrice[1] * float(getStartRoute[2])), 2))

                ws.cell(row=startRow, column=startColumn+7, value=petrolPrice).alignment = Alignment(vertical="center", horizontal="center")  # petrol price
                ws.cell(row=startRow, column=startColumn+8, value=diets).alignment = Alignment(vertical="center", horizontal="center")  # diets
                ws.cell(row=startRow, column=startColumn+11, value=petrolPrice+float(diets)).alignment = Alignment(vertical="center", horizontal="center")  # together
            if event == 1:
                ws.cell(row=startRow, column=startColumn+1, value="príchod")  # fill prichod
                ws.cell(row=startRow, column=startColumn+2, value=getStartRoute[1])  # fill destination city
                ws.cell(row=startRow, column=startColumn+3, value=getRandomTime[1]).alignment = Alignment(vertical="center", horizontal="center")  # morning end time

            # way back <-
            if event == 2:
                ws.cell(row=startRow, column=startColumn+1, value="odchod")  # fill odchod
                ws.cell(row=startRow, column=startColumn+2, value=getStartRoute[1])  # fill destination city
                ws.cell(row=startRow, column=startColumn+3, value=getRandomTime[2]).alignment = Alignment(vertical="center", horizontal="center")  # evening start time
                ws.cell(row=startRow, column=startColumn+4, value="AUS").alignment = Alignment(vertical="center", horizontal="center")
                ws.cell(row=startRow, column=startColumn+5, value=getStartRoute[2]).alignment = Alignment(vertical="center", horizontal="center")  # km
                ws.cell(row=startRow, column=startColumn+7, value=petrolPrice).alignment = Alignment(vertical="center", horizontal="center")  # petrol price
                ws.cell(row=startRow, column=startColumn+8, value=diets).alignment = Alignment(vertical="center", horizontal="center")  # diets
                ws.cell(row=startRow, column=startColumn+11, value=petrolPrice+float(diets)).alignment = Alignment(vertical="center", horizontal="center")  # together
            if event == 3:
                ws.cell(row=startRow, column=startColumn+1, value="príchod")  # fill prichod
                ws.cell(row=startRow, column=startColumn+2, value=getStartRoute[0])  # fill start city
                ws.cell(row=startRow, column=startColumn+3, value=getRandomTime[3]).alignment = Alignment(vertical="center", horizontal="center")  # evening end time

            startRow += 1
        
        # increment day by 1
        startDate = startDate + datetime.timedelta(days=1)

    writeFooterValues()

# main
if __name__ == '__main__':
    print("Processing output.xlsx ...")
    wb1 = load_workbook("mesta_input.xlsx")
    wb2 = load_workbook("output.xlsx")
    ws = wb2.active  # set 2nd excel active

    # generateDates params: startRow, startColumn, startDate, numberOfDays
    fillSheet(5, 1, datetime.datetime.strptime('2020-01-01', '%Y-%m-%d'), 31)
    wb2.save("output.xlsx")
    print("Done!")
